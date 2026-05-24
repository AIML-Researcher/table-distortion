"""
Distortion Reviewer
===================
Validates that the distortions produced by ``Distortion_Generator`` are
actually effective.  For every distortion sheet inside a generated
``.xlsx`` workbook the reviewer uses an LLM agent to:

  1. **Direct-answer attempt** – try to answer the original question on
     the *distorted* table as-is.  A "good" distortion should make this
     fail (the answer should not match the ground-truth).
  2. **De-distortion + answer** – only when step 1 has confirmed that
     the distortion blocks direct answering, ask the agent to detect
     the distortion, reverse it ("de-distort"), and then answer the
     original question on the reconstructed table.  A "good" distortion
     must still be reversible, so this second attempt should recover
     the correct answer.

A distortion is considered **validated** when:

  * direct answer  →  INCORRECT (blocks direct QA),            AND
  * de-distorted   →  CORRECT   (information preserved / recoverable).

Any other combination is flagged (distortion too weak, or destructive).

The module runs independently - no argparse / CLI; the entry-points are
plain Python functions so they can be called from notebooks or from
``Run_Pipelines.py`` just like ``Distortion_Selector`` and
``Distortion_Generator``.

Usage (library)
---------------
    from Distortion_Reviewer import run_reviewer

    run_reviewer(
        samples_json="30_Samples.json",
        distorted_dir="Output/04162026",
        output_json="30_Samples_reviewed.json",
    )
"""


from __future__ import annotations
import json
import os
import re
from pathlib import Path
from typing import Any
import pandas as pd
import shutil
from openpyxl import load_workbook
from azure.identity import DefaultAzureCredential, get_bearer_token_provider
from openai import AzureOpenAI

from Python_Sandbox import Sandbox, ExecutionResult


# ---------------------------------------------------------------------------
# Sandbox singleton (lazy) – the reviewer lets the LLM generate Python code
# that it then runs in an isolated subprocess to produce an answer.
# ---------------------------------------------------------------------------

_SANDBOX: Sandbox | None = None
_SANDBOX_READY: bool = False


def _get_sandbox() -> Sandbox:
    """Return a lazily-initialised sandbox with pandas/openpyxl installed."""
    global _SANDBOX, _SANDBOX_READY
    if _SANDBOX is None:
        _SANDBOX = Sandbox()
    sb = _SANDBOX
    if not _SANDBOX_READY:
        # Install the libraries the generated code is most likely to need.
        # pip is a no-op if they are already present.
        try:
            r = sb.pip_install(
                ["pandas", "openpyxl", "numpy"], timeout=600.0
            )
            if not r.ok:
                print("[reviewer] ⚠ sandbox pip_install returned "
                      f"rc={r.returncode}: {r.stderr[:400]}")
        except Exception as e:
            print(f"[reviewer] ⚠ sandbox pip_install failed: {e}")
        _SANDBOX_READY = True
    return sb


# ---------------------------------------------------------------------------
# Workbook helpers
# ---------------------------------------------------------------------------

ORIGINAL_SHEET_NAME = "original"


# ---------------------------------------------------------------------------
# Distortion catalogue – GENERAL descriptions of what each distortion does
# (not the specific parameters used on any given table).  Used to give the
# phase-2 agent a hint about what kind of corruption it is expected to
# reverse, without revealing the exact transformation applied.
# ---------------------------------------------------------------------------

DISTORTION_CATALOG: dict[str, str] = {
    "merge_cells":
        "Structural distortion where some adjacent cells (typically in the "
        "header row or along a column) are merged into a single cell, so "
        "one cell visually spans multiple rows/columns. In the serialised "
        "sheet the non-anchor cells of the merge appear empty while the "
        "anchor cell holds the value.",
    "horizontal_shift":
        "Structural distortion where the values in a subset of rows are "
        "shifted left or right by one or more columns, so data ends up "
        "under the wrong headers for those rows.",
    "vertical_shift":
        "Structural distortion where one or more entire columns (header "
        "and all their values) are shifted DOWN by 1-4 rows, leaving the "
        "top cells of those columns empty and breaking the row-level "
        "alignment between the shifted columns and the rest of the table.",
    "broken_rows_split":
        "Structural distortion where some rows of the original table are "
        "split vertically into two consecutive rows, each holding a "
        "portion of the original row's cells (the rest being empty).",
    "broken_rows_merge":
        "Structural distortion where two consecutive original rows are "
        "merged into a single row by concatenating their cell values "
        "(usually with a space or newline) into the same columns.",
    "ocr_char_misinterpretation":
        "OCR-style distortion where individual characters are substituted "
        "with visually similar ones (0↔O, 1↔l↔I, 5↔S, 8↔B, 2↔Z, …) in "
        "random cells.",
    "ocr_lost_formatting":
        "OCR-style distortion where formatting characters and symbols "
        "such as %, $, superscripts/subscripts, thousands separators, "
        "currency symbols, etc. are stripped or mangled inside cell "
        "values.",
    "numbers_as_text":
        "Data-type distortion where numeric values are stored as strings, "
        "often with thousands separators or surrounding whitespace/quotes, "
        "so they no longer parse as numbers.",
    "date_format_corruption":
        "Data-type distortion where date cells are reformatted into an "
        "inconsistent / ambiguous format (e.g. swapping day and month, "
        "changing separators, mixing styles across the column).",
    "decimal_separator_swap":
        "Data-type distortion where decimal and thousands separators are "
        "swapped (European vs US convention), e.g. '1,234.56' ↔ "
        "'1.234,56', causing numbers to parse incorrectly.",
    "header_as_data_row":
        "Schema distortion where the original header row is demoted into "
        "the first data row and a generic header (or the first data row) "
        "is promoted into the header position.",
    "multi_column_collapse":
        "Layout distortion where two or more adjacent columns are "
        "collapsed into one, concatenating their values (with a space or "
        "delimiter) and removing the original column boundary.",
    "footnote_injection":
        "Layout distortion where one or more extra rows containing "
        "footnote / side-note text (not part of the tabular data) are "
        "inserted, typically at the bottom of the table.",
    "random_noise_characters":
        "OCR-style distortion where random noise characters (e.g. "
        "#@?~^&*!`|) are sprinkled into existing cell values at random "
        "positions.",
    "context_loss":
        "Semantic distortion where units, currency symbols, or other "
        "contextual suffixes/prefixes (kg, km, $, %, …) are stripped from "
        "cell values, leaving bare numbers without their context.",
}


def _describe_distortion(name: str) -> str:
    """Return a generic description of a distortion.

    Falls back to a vague placeholder if the distortion name is unknown
    (e.g. the sheet was produced by a newer / custom distortion)."""
    if name in DISTORTION_CATALOG:
        return DISTORTION_CATALOG[name]
    return (
        "An unspecified distortion from the known catalogue (structural, "
        "OCR, data-type, schema, layout, or semantic). Inspect the sheet "
        "to infer the specific kind."
    )


def _sheet_to_dataframe(wb, sheet_name: str) -> pd.DataFrame:
    """Read an openpyxl sheet into a DataFrame (all strings)."""
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(["" if v is None else str(v) for v in row])

    if not rows:
        return pd.DataFrame()

    header = rows[0]
    data = rows[1:]
    # Make header unique to avoid pandas issues with duplicates
    seen: dict[str, int] = {}
    uniq_header = []
    for h in header:
        if h in seen:
            seen[h] += 1
            uniq_header.append(f"{h}.{seen[h]}")
        else:
            seen[h] = 0
            uniq_header.append(h)
    return pd.DataFrame(data, columns=uniq_header)


def _df_to_markdown(df: pd.DataFrame, max_rows: int = 80) -> str:
    """Serialise a DataFrame to a compact Markdown table string."""
    if df.empty:
        return "(empty table)"
    if len(df) > max_rows:
        top = df.head(max_rows // 2)
        bot = df.tail(max_rows // 2)
        ellipsis_row = pd.DataFrame(
            [["…"] * len(df.columns)], columns=df.columns
        )
        df = pd.concat([top, ellipsis_row, bot], ignore_index=True)
    try:
        return df.to_markdown(index=False)
    except Exception:
        return df.to_csv(index=False)


# ---------------------------------------------------------------------------
# Answer-comparison helper
# ---------------------------------------------------------------------------


_PUNCT_RE = re.compile(r"[^\w\s]")


def _normalise(s: str) -> str:
    s = str(s).lower().strip()
    s = _PUNCT_RE.sub(" ", s)
    s = re.sub(r"\s+", " ", s)
    return s


def _answers_match(predicted: str, gold: str) -> bool:
    """Loose equality: ignore punctuation/case/whitespace; also accept
    either string being a substring of the other."""
    if predicted is None:
        return False
    p, g = _normalise(predicted), _normalise(gold)
    if not p or not g:
        return False
    if p == g:
        return True
    return (p in g) or (g in p)


# ---------------------------------------------------------------------------
# LLM plumbing
# ---------------------------------------------------------------------------


def _make_client() -> tuple[AzureOpenAI, str]:
    endpoint = os.getenv(
        "ENDPOINT_URL",
        "https://oaiscience-oai-swc.openai.azure.com/",
    )
    deployment = os.getenv("DEPLOYMENT_NAME", "gpt-4o-mini")
    token_provider = get_bearer_token_provider(
        DefaultAzureCredential(),
        "https://cognitiveservices.azure.com/.default",
    )
    client = AzureOpenAI(
        azure_endpoint=endpoint,
        azure_ad_token_provider=token_provider,
        api_version="2025-01-01-preview",
    )
    return client, deployment


def _chat_json(client: AzureOpenAI, deployment: str,
               system: str, user: str,
               model: str | None = None,
               temperature: float = 0.0) -> dict[str, Any]:
    response = client.chat.completions.create(
        model=model or deployment,
        temperature=temperature,
        messages=[
            {"role": "system", "content": system},
            {"role": "user", "content": user},
        ],
        response_format={"type": "json_object"},
    )
    raw = response.choices[0].message.content.strip()
    if raw.startswith("```"):
        raw = re.sub(r"^```(?:json)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
    return json.loads(raw)


# ---------------------------------------------------------------------------
# Prompts
# ---------------------------------------------------------------------------


DIRECT_ANSWER_PROMPT = """\
You are a meticulous Table-QA agent that answers questions by WRITING
AND EXECUTING Python code.

You will receive:
  • The ABSOLUTE path to an .xlsx workbook on disk (``XLSX_PATH``).
  • The exact SHEET NAME inside that workbook holding a table that MAY
    be corrupted / distorted in subtle ways.
  • A Markdown preview of that sheet so you can inspect its structure.
  • A natural-language question.

Your task:
  Produce Python code that, when executed, answers the question using
  ONLY what can be read directly from the sheet in its CURRENT
  (possibly distorted) form.  Do NOT try to detect, repair, or reverse
  the distortion – treat the table as-is.  If the distortion makes it
  impossible to answer reliably, still produce your best-effort answer
  (or an empty answer string) and lower the confidence.

Rules for the generated code:
  • It must be self-contained, run under Python 3, and read the xlsx
    from ``XLSX_PATH`` (available as an environment variable and as
    ``sys.argv[1]``; the sheet name is in ``SHEET_NAME`` /
    ``sys.argv[2]``).
  • You may use pandas, openpyxl, numpy, stdlib – they are installed.
  • The VERY LAST line written to stdout MUST be a single JSON object
    on ONE line with this schema (no trailing text afterwards):
        {"answer": "<string>", "reasoning": "<short>",
         "confidence": "high" | "medium" | "low",
         "answerable_directly": true | false}
  • Do NOT print anything after that final JSON line.
  • The code must not require network access or user input.

Respond ONLY with a JSON object with this exact schema:
```json
{
  "plan": "<one-paragraph description of how your code answers the question>",
  "code": "<the full Python source code as a string>"
}
```
Return ONLY that JSON, no extra commentary.
"""


DEDISTORT_ANSWER_PROMPT = """\
You are an expert Table-QA agent specialised in DETECTING and REVERSING
(de-distorting) corruption in tables before answering questions – by
WRITING AND EXECUTING Python code.

You will receive:
  • The ABSOLUTE path to an .xlsx workbook on disk (``XLSX_PATH``).
  • The SHEET NAME of a table that has been corrupted.
  • The NAME of the distortion category that was applied to this sheet,
    together with a GENERAL description of what that distortion does.
    NOTE: you are told the *kind* of distortion, but NOT the exact
    parameters used on this specific table (e.g. which rows were
    shifted, by how many columns, which characters were swapped, etc.).
    You must still inspect the sheet to figure out the specifics.
  • A Markdown preview of that sheet.
  • The original natural-language question.

Your task:
  1. Using the provided distortion description as a strong prior,
     inspect the sheet and determine the specific way the distortion
     was applied to this table.
  2. Write Python code that programmatically reconstructs ("de-distorts")
     the table to its likely pre-distortion form.
  3. In the same code, answer the question on the reconstructed table.

Rules for the generated code:
  • Self-contained Python 3.  Read the xlsx from ``XLSX_PATH``
    (available as env var and as ``sys.argv[1]``; sheet name in
    ``SHEET_NAME`` / ``sys.argv[2]``).
  • You may use pandas, openpyxl, numpy, stdlib.
  • The VERY LAST line on stdout MUST be a single JSON object on ONE
    line with this schema (no trailing text):
        {"answer": "<string>",
         "detected_distortion": "<short name>",
         "reconstruction_notes": "<brief>",
         "confidence": "high" | "medium" | "low"}
  • Do NOT print anything after that final JSON line.
  • No network / user input.

Respond ONLY with a JSON object:
```json
{
  "plan": "<paragraph describing your detection + reconstruction strategy>",
  "code": "<the full Python source code as a string>"
}
```
Return ONLY that JSON.
"""


# ---------------------------------------------------------------------------
# Code-execution helper
# ---------------------------------------------------------------------------


_JSON_LINE_RE = re.compile(r"\{.*\}\s*$", re.DOTALL)


def _extract_final_json(stdout: str) -> dict[str, Any] | None:
    """Return the last JSON object printed to stdout, or None."""
    if not stdout:
        return None
    # Try last non-empty line first.
    lines = [ln for ln in stdout.splitlines() if ln.strip()]
    for ln in reversed(lines):
        s = ln.strip()
        if s.startswith("{") and s.endswith("}"):
            try:
                return json.loads(s)
            except Exception:
                pass
    # Fallback: greedy match of the last {...} block in the whole stdout.
    m = _JSON_LINE_RE.search(stdout.strip())
    if m:
        try:
            return json.loads(m.group(0))
        except Exception:
            return None
    return None


def _run_generated_code(code: str,
                        xlsx_path: str,
                        sheet_name: str,
                        timeout: float = 60.0) -> tuple[ExecutionResult,
                                                        dict[str, Any] | None]:
    """Execute LLM-generated code in the sandbox and parse its final JSON."""
    sb = _get_sandbox()
    abs_xlsx = str(Path(xlsx_path).resolve())
    result = sb.execute(
        code,
        timeout=timeout,
        extra_env={"XLSX_PATH": abs_xlsx, "SHEET_NAME": sheet_name},
        argv=[abs_xlsx, sheet_name],
    )
    parsed = _extract_final_json(result.stdout) if result.ok else None
    return result, parsed


def _ask_for_code(client: AzureOpenAI, deployment: str,
                  system: str, user: str,
                  model: str | None) -> dict[str, Any]:
    """Ask the LLM for {plan, code}."""
    return _chat_json(client, deployment, system=system, user=user, model=model)


# ---------------------------------------------------------------------------
# Core per-distortion review
# ---------------------------------------------------------------------------


def _review_one_distortion(client: AzureOpenAI, deployment: str,
                           distortion_name: str,
                           xlsx_path: str,
                           distorted_md: str,
                           question: str,
                           gold_answer: str,
                           model: str | None = None,
                           exec_timeout: float = 60.0) -> dict[str, Any]:
    """Run the two-phase code-generating review for a single distortion sheet."""

    user_msg = (
        f"### XLSX_PATH\n{Path(xlsx_path).resolve()}\n\n"
        f"### SHEET_NAME\n{distortion_name}\n\n"
        f"### Distorted sheet preview (Markdown)\n{distorted_md}\n\n"
        f"### Question\n{question}"
    )

    # Phase-2 gets an extra hint: the distortion category name + a generic
    # description of what that kind of distortion does in general, but NOT
    # the exact parameters used on this particular table.
    distortion_desc = _describe_distortion(distortion_name)
    user_msg_phase2 = (
        f"### XLSX_PATH\n{Path(xlsx_path).resolve()}\n\n"
        f"### SHEET_NAME\n{distortion_name}\n\n"
        f"### Applied distortion (name)\n{distortion_name}\n\n"
        f"### Applied distortion (general description)\n{distortion_desc}\n\n"
        f"### Distorted sheet preview (Markdown)\n{distorted_md}\n\n"
        f"### Question\n{question}"
    )

    result: dict[str, Any] = {"distortion": distortion_name}

    # --- Phase 1: generate code that directly answers the question -------
    try:
        gen1 = _ask_for_code(
            client, deployment,
            system=DIRECT_ANSWER_PROMPT,
            user=user_msg,
            model=model,
        )
    except Exception as e:
        return {
            **result,
            "error": f"direct-answer code-gen call failed: {e}",
            "validated": False,
        }

    code1 = str(gen1.get("code", ""))
    plan1 = gen1.get("plan")
    exec1, parsed1 = _run_generated_code(
        code1, xlsx_path, distortion_name, timeout=exec_timeout
    )
    direct_answer = ""
    direct_conf = None
    direct_answerable = None
    direct_reasoning = None
    if parsed1 is not None:
        direct_answer = str(parsed1.get("answer", ""))
        direct_conf = parsed1.get("confidence")
        direct_answerable = parsed1.get("answerable_directly")
        direct_reasoning = parsed1.get("reasoning")

    direct_correct = _answers_match(direct_answer, gold_answer)
    result["direct"] = {
        "plan": plan1,
        "code": code1,
        "stdout": exec1.stdout,
        "stderr": exec1.stderr,
        "returncode": exec1.returncode,
        "timed_out": exec1.timed_out,
        "duration": exec1.duration,
        "parsed": parsed1,
        "answer": direct_answer,
        "correct": direct_correct,
        "confidence": direct_conf,
        "answerable_directly": direct_answerable,
        "reasoning": direct_reasoning,
    }

    if direct_correct:
        result["validated"] = False
        result["verdict"] = "ineffective"
        result["explanation"] = (
            "Generated code could answer the question directly from the "
            "distorted sheet – distortion did not block direct QA."
        )
        return result

    # --- Phase 2: generate code that de-distorts then answers ------------
    try:
        gen2 = _ask_for_code(
            client, deployment,
            system=DEDISTORT_ANSWER_PROMPT,
            user=user_msg_phase2,
            model=model,
        )
    except Exception as e:
        result["error"] = f"de-distort code-gen call failed: {e}"
        result["validated"] = False
        return result

    code2 = str(gen2.get("code", ""))
    plan2 = gen2.get("plan")
    exec2, parsed2 = _run_generated_code(
        code2, xlsx_path, distortion_name, timeout=exec_timeout
    )
    dd_answer = ""
    dd_conf = None
    dd_detected = None
    dd_notes = None
    if parsed2 is not None:
        dd_answer = str(parsed2.get("answer", ""))
        dd_conf = parsed2.get("confidence")
        dd_detected = parsed2.get("detected_distortion")
        dd_notes = parsed2.get("reconstruction_notes")

    dd_correct = _answers_match(dd_answer, gold_answer)
    result["dedistort"] = {
        "plan": plan2,
        "code": code2,
        "stdout": exec2.stdout,
        "stderr": exec2.stderr,
        "returncode": exec2.returncode,
        "timed_out": exec2.timed_out,
        "duration": exec2.duration,
        "parsed": parsed2,
        "answer": dd_answer,
        "correct": dd_correct,
        "confidence": dd_conf,
        "detected_distortion": dd_detected,
        "reconstruction_notes": dd_notes,
    }

    if dd_correct:
        result["validated"] = True
        result["verdict"] = "validated"
        result["explanation"] = (
            "Direct-answer code was wrong, but de-distorting code recovered "
            "the correct answer."
        )
    else:
        result["validated"] = False
        result["verdict"] = "destructive"
        result["explanation"] = (
            "Even after de-distortion code execution, the agent could not "
            "recover the correct answer – distortion may be irreversible "
            "or the generated code was inadequate."
        )

    return result


# ---------------------------------------------------------------------------
# Public API – per-sample review
# ---------------------------------------------------------------------------


def review_distorted_workbook(xlsx_path: str,
                              question: str,
                              gold_answer: str,
                              distortions: list[str] | None = None,
                              model: str | None = None,
                              verbose: bool = True) -> dict[str, Any]:
    """Review every distortion sheet inside ``xlsx_path``."""
    wb = load_workbook(xlsx_path, data_only=True)
    sheet_names = wb.sheetnames

    if distortions is None:
        distortions = [s for s in sheet_names if s != ORIGINAL_SHEET_NAME]
    else:
        distortions = [d for d in distortions if d in sheet_names]

    client, deployment = _make_client()

    results: list[dict[str, Any]] = []
    for name in distortions:
        if verbose:
            print(f"    • Reviewing distortion: {name}")
        df_dist = _sheet_to_dataframe(wb, name)
        md_dist = _df_to_markdown(df_dist)
        res = _review_one_distortion(
            client=client,
            deployment=deployment,
            distortion_name=name,
            xlsx_path=xlsx_path,
            distorted_md=md_dist,
            question=question,
            gold_answer=gold_answer,
            model=model,
        )
        if verbose:
            print(f"        → {res.get('verdict', 'error')}")
        results.append(res)

    validated = [r["distortion"] for r in results if r.get("validated")]
    summary = {
        "total": len(results),
        "validated": len(validated),
        "ineffective": sum(1 for r in results
                           if r.get("verdict") == "ineffective"),
        "destructive": sum(1 for r in results
                           if r.get("verdict") == "destructive"),
        "errors": sum(1 for r in results if "error" in r),
        "validated_distortions": validated,
    }

    return {
        "xlsx": str(xlsx_path),
        "question": question,
        "answer": gold_answer,
        "results": results,
        "summary": summary,
    }


# ---------------------------------------------------------------------------
# Public API – batch review driven by a samples JSON
# ---------------------------------------------------------------------------


def _infer_xlsx_path(distorted_dir: str, sample: dict[str, Any]) -> Path:
    """Given a samples entry like ``{"filepath": "csv/204-csv/590.csv"}``
    return the matching ``<distorted_dir>/204-csv/590.xlsx`` path."""
    rel = sample["filepath"]
    if rel.startswith("csv/") or rel.startswith("csv\\"):
        rel = rel[4:]
    folder, file = rel.replace("\\", "/").split("/")
    stem = Path(file).stem
    return Path(distorted_dir) / folder / f"{stem}.xlsx"


def _write_validated_workbook(src_xlsx: Path,
                              dst_xlsx: Path,
                              validated_sheets: list[str],
                              store_original: bool = True,
                              verbose: bool = True) -> None:
    """Copy ``src_xlsx`` to ``dst_xlsx`` keeping only the sheets listed in
    ``validated_sheets`` (non-validated sheets are removed).

    When ``store_original`` is True, the sheet containing the original
    (undistorted) table – ``ORIGINAL_SHEET_NAME`` – is also retained and
    moved to the first position of the workbook.
    """
    dst_xlsx.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(src_xlsx, dst_xlsx)

    wb = load_workbook(dst_xlsx)
    keep = set(validated_sheets)
    if store_original and ORIGINAL_SHEET_NAME in wb.sheetnames:
        keep.add(ORIGINAL_SHEET_NAME)

    # openpyxl requires at least one sheet at all times; drop others last.
    to_remove = [s for s in wb.sheetnames if s not in keep]
    # Ensure we don't remove everything (shouldn't happen – caller checks).
    if len(to_remove) == len(wb.sheetnames):
        if verbose:
            print(f"      ⚠ No validated sheets to keep in {src_xlsx.name}")
        return
    for s in to_remove:
        del wb[s]

    # Move the original sheet to the first position if requested.
    if store_original and ORIGINAL_SHEET_NAME in wb.sheetnames:
        idx = wb.sheetnames.index(ORIGINAL_SHEET_NAME)
        if idx != 0:
            wb.move_sheet(ORIGINAL_SHEET_NAME, offset=-idx)

    wb.save(dst_xlsx)
    if verbose:
        print(f"      ✓ Stored validated workbook → {dst_xlsx} "
              f"(sheets: {wb.sheetnames})")


def run_reviewer(samples_json: str,
                 distorted_dir: str | None = None,
                 output_json: str | None = None,
                 stats_json: str | None = None,
                 limit: int | None = None,
                 model: str | None = None,
                 store_validated: bool = False,
                 verbose: bool = True) -> list[dict[str, Any]]:
    """Batch-run the reviewer over every sample in ``samples_json``.

    Supports two JSON formats:

    **Legacy format** (used with ``distorted_dir``):
        Each entry has ``filepath``, ``question``, ``answer``, and
        optionally ``suitable_distortions``.  The workbook path is
        inferred via ``_infer_xlsx_path``.

    **Generator format** (from ``Create_Generator_JSON``):
        Each entry has ``data_file`` (path to a single-distortion xlsx),
        ``original_file``, ``query``, ``answer``, ``distortion``,
        ``distortion_type``, ``index``, ``dtype``.  No ``distorted_dir``
        is needed since paths are explicit.

    The output JSON (``output_json``) is written in the **same format**
    as the input, with each entry augmented with:
      - ``phase1_passed`` (bool): direct QA on distorted table failed
      - ``phase2_passed`` (bool): de-distortion recovered correct answer
      - ``verdict``: "validated" | "ineffective" | "destructive" | "error"

    Aggregate statistics are written to a separate ``stats_json`` file.

    When ``store_validated`` is True and ``distorted_dir`` is provided
    (legacy mode), validated workbooks are copied to a sibling folder.
    """
    with open(samples_json, "r", encoding="utf-8") as fr:
        samples = json.load(fr)

    if limit is not None:
        samples = samples[:limit]

    # Detect format: generator format has "data_file" key
    is_generator_format = (
        len(samples) > 0 and "data_file" in samples[0]
    )

    distorted_root: Path | None = None
    validated_root: Path | None = None
    if distorted_dir:
        distorted_root = Path(distorted_dir)
        if store_validated:
            validated_root = (
                distorted_root.parent / f"validated_{distorted_root.name}"
            )
            validated_root.mkdir(parents=True, exist_ok=True)
            if verbose:
                print(f"📁 Storing validated workbooks under {validated_root}")

    reviewed: list[dict[str, Any]] = []
    # Aggregate counters
    dist_phase1_pass = 0
    dist_phase1_fail = 0
    dist_phase2_pass = 0
    dist_phase2_fail = 0
    dist_errors = 0

    # Per-distortion-type breakdown
    type_stats: dict[str, dict[str, int]] = {}
    # Per-distortion-name breakdown
    distortion_stats: dict[str, dict[str, int]] = {}

    client, deployment = _make_client()

    for n, sample in enumerate(samples):
        if is_generator_format:
            # --- Generator format: each entry is one distortion in its own xlsx ---
            xlsx_path = Path(sample["data_file"])
            question = sample["query"]
            gold_answer = sample["answer"]
            distortion_name = sample["distortion"]
            distortion_type = sample.get("distortion_type", "Unknown")

            if verbose:
                print(f"{n}. Reviewing {xlsx_path} ({distortion_name}) ...")

            if not xlsx_path.exists():
                if verbose:
                    print(f"   ⚠ Missing file, skipped: {xlsx_path}")
                sample_out = dict(sample)
                sample_out["phase1_passed"] = None
                sample_out["phase2_passed"] = None
                sample_out["verdict"] = "error"
                sample_out["error"] = f"missing file: {xlsx_path}"
                reviewed.append(sample_out)
                dist_errors += 1
                continue

            # Load the workbook – the distorted data is in the first sheet
            try:
                wb = load_workbook(xlsx_path, data_only=True)
                # Use the first sheet (the distortion sheet)
                sheet_name = wb.sheetnames[0]
                df_dist = _sheet_to_dataframe(wb, sheet_name)
                md_dist = _df_to_markdown(df_dist)

                res = _review_one_distortion(
                    client=client,
                    deployment=deployment,
                    distortion_name=sheet_name,
                    xlsx_path=str(xlsx_path),
                    distorted_md=md_dist,
                    question=question,
                    gold_answer=gold_answer,
                    model=model,
                )
            except Exception as e:
                if verbose:
                    print(f"   ⚠ Review failed: {e}")
                sample_out = dict(sample)
                sample_out["phase1_passed"] = None
                sample_out["phase2_passed"] = None
                sample_out["verdict"] = "error"
                sample_out["error"] = str(e)
                reviewed.append(sample_out)
                dist_errors += 1
                continue

            # Determine phase1/phase2 outcomes
            direct = res.get("direct") or {}
            dd = res.get("dedistort")
            phase1_passed = direct.get("correct") is False
            phase2_passed = (dd is not None and dd.get("correct") is True) if phase1_passed else False

            sample_out = dict(sample)
            sample_out["phase1_passed"] = phase1_passed
            sample_out["phase2_passed"] = phase2_passed
            sample_out["verdict"] = res.get("verdict", "error")
            reviewed.append(sample_out)

            # Update counters
            if direct.get("correct") is True:
                dist_phase1_fail += 1
            elif direct.get("correct") is False:
                dist_phase1_pass += 1
                if dd is None:
                    dist_errors += 1
                elif dd.get("correct") is True:
                    dist_phase2_pass += 1
                else:
                    dist_phase2_fail += 1
            else:
                dist_errors += 1

            # Per-type stats
            if distortion_type not in type_stats:
                type_stats[distortion_type] = {
                    "total": 0, "phase1_pass": 0, "phase1_fail": 0,
                    "phase2_pass": 0, "phase2_fail": 0, "errors": 0,
                }
            ts = type_stats[distortion_type]
            ts["total"] += 1
            if phase1_passed:
                ts["phase1_pass"] += 1
                if phase2_passed:
                    ts["phase2_pass"] += 1
                elif dd is not None:
                    ts["phase2_fail"] += 1
                else:
                    ts["errors"] += 1
            elif direct.get("correct") is True:
                ts["phase1_fail"] += 1
            else:
                ts["errors"] += 1

            # Per-distortion-name stats
            if distortion_name not in distortion_stats:
                distortion_stats[distortion_name] = {
                    "total": 0, "phase1_pass": 0, "phase1_fail": 0,
                    "phase2_pass": 0, "phase2_fail": 0, "errors": 0,
                }
            ds = distortion_stats[distortion_name]
            ds["total"] += 1
            if phase1_passed:
                ds["phase1_pass"] += 1
                if phase2_passed:
                    ds["phase2_pass"] += 1
                elif dd is not None:
                    ds["phase2_fail"] += 1
                else:
                    ds["errors"] += 1
            elif direct.get("correct") is True:
                ds["phase1_fail"] += 1
            else:
                ds["errors"] += 1

            if verbose:
                print(f"        → {sample_out['verdict']}")

        else:
            # --- Legacy format ---
            xlsx_path = _infer_xlsx_path(distorted_dir, sample)
            if verbose:
                print(f"{n}. Reviewing {xlsx_path} ...")
            if not xlsx_path.exists():
                if verbose:
                    print(f"   ⚠ Missing workbook, skipped: {xlsx_path}")
                sample_out = dict(sample)
                sample_out["phase1_passed"] = None
                sample_out["phase2_passed"] = None
                sample_out["verdict"] = "error"
                sample_out["error"] = f"missing workbook: {xlsx_path}"
                reviewed.append(sample_out)
                continue

            try:
                review = review_distorted_workbook(
                    xlsx_path=str(xlsx_path),
                    question=sample["question"],
                    gold_answer=sample["answer"],
                    distortions=sample.get("suitable_distortions"),
                    model=model,
                    verbose=verbose,
                )
            except Exception as e:
                if verbose:
                    print(f"   ⚠ Review failed: {e}")
                sample_out = dict(sample)
                sample_out["phase1_passed"] = None
                sample_out["phase2_passed"] = None
                sample_out["verdict"] = "error"
                sample_out["error"] = str(e)
                reviewed.append(sample_out)
                continue

            # Compute phase1/phase2 from review results
            any_p1_pass = False
            any_p2_pass = False
            for r in review.get("results", []):
                if "error" in r and "direct" not in r:
                    dist_errors += 1
                    continue
                direct = r.get("direct") or {}
                if direct.get("correct") is True:
                    dist_phase1_fail += 1
                elif direct.get("correct") is False:
                    dist_phase1_pass += 1
                    any_p1_pass = True
                    dd = r.get("dedistort")
                    if dd is None:
                        dist_errors += 1
                    elif dd.get("correct") is True:
                        dist_phase2_pass += 1
                        any_p2_pass = True
                    else:
                        dist_phase2_fail += 1
                else:
                    dist_errors += 1

            sample_out = dict(sample)
            sample_out["phase1_passed"] = any_p1_pass
            sample_out["phase2_passed"] = any_p2_pass
            sample_out["verdict"] = "validated" if any_p2_pass else (
                "ineffective" if not any_p1_pass else "destructive"
            )
            sample_out["validated_distortions"] = review.get("summary", {}).get(
                "validated_distortions", []
            )

            if store_validated and validated_root is not None:
                validated_list = sample_out.get("validated_distortions", [])
                if validated_list:
                    try:
                        rel = xlsx_path.relative_to(distorted_root)
                    except ValueError:
                        rel = Path(xlsx_path.parent.name) / xlsx_path.name
                    dst = validated_root / rel
                    try:
                        _write_validated_workbook(
                            src_xlsx=xlsx_path,
                            dst_xlsx=dst,
                            validated_sheets=validated_list,
                            verbose=verbose,
                        )
                    except Exception as e:
                        if verbose:
                            print(f"   ⚠ Failed to store validated workbook: {e}")

            reviewed.append(sample_out)

    # -------- Aggregate statistics -------------------------------------
    def _pct(num: int, denom: int) -> float:
        return round(100.0 * num / denom, 2) if denom else 0.0

    dist_total = dist_phase1_pass + dist_phase1_fail + dist_errors
    dist_phase2_total = dist_phase2_pass + dist_phase2_fail
    stats = {
        "distortions": {
            "total": dist_total,
            "phase1_pass": dist_phase1_pass,
            "phase1_pass_pct": _pct(dist_phase1_pass, dist_total),
            "phase1_fail": dist_phase1_fail,
            "phase1_fail_pct": _pct(dist_phase1_fail, dist_total),
            "phase2_pass": dist_phase2_pass,
            "phase2_pass_pct": _pct(dist_phase2_pass, dist_phase2_total),
            "phase2_fail": dist_phase2_fail,
            "phase2_fail_pct": _pct(dist_phase2_fail, dist_phase2_total),
            "errors": dist_errors,
            "errors_pct": _pct(dist_errors, dist_total),
        },
        "by_distortion_type": {
            dtype: {
                **counts,
                "phase1_pass_pct": _pct(counts["phase1_pass"], counts["total"]),
                "phase2_pass_pct": _pct(counts["phase2_pass"],
                                        counts["phase2_pass"] + counts["phase2_fail"]),
            }
            for dtype, counts in type_stats.items()
        },
        "distortion_wise": {
            dname: {
                **counts,
                "phase1_pass_pct": _pct(counts["phase1_pass"], counts["total"]),
                "phase2_pass_pct": _pct(counts["phase2_pass"],
                                        counts["phase2_pass"] + counts["phase2_fail"]),
            }
            for dname, counts in distortion_stats.items()
        },
    }

    if verbose:
        d = stats["distortions"]
        print("\n" + "=" * 72)
        print("REVIEW STATISTICS")
        print("=" * 72)
        print(f"Distortions reviewed: {d['total']}")
        print(f"  Phase 1 (direct QA blocked)  "
              f"pass: {d['phase1_pass']:4d} ({d['phase1_pass_pct']:.2f}%)  "
              f"fail: {d['phase1_fail']:4d} ({d['phase1_fail_pct']:.2f}%)")
        print(f"  Phase 2 (reversible / valid) "
              f"pass: {d['phase2_pass']:4d} ({d['phase2_pass_pct']:.2f}%)  "
              f"fail: {d['phase2_fail']:4d} ({d['phase2_fail_pct']:.2f}%)   "
              f"[of {dist_phase2_total} that passed phase 1]")
        print(f"  Errors: {d['errors']} ({d['errors_pct']:.2f}%)")
        if type_stats:
            print(f"\n  By distortion type:")
            for dtype, counts in type_stats.items():
                print(f"    {dtype}: total={counts['total']}  "
                      f"p1_pass={counts['phase1_pass']}  "
                      f"p2_pass={counts['phase2_pass']}")
        if distortion_stats:
            print(f"\n  By distortion (distortion_wise):")
            for dname, counts in distortion_stats.items():
                print(f"    {dname}: total={counts['total']}  "
                      f"p1_pass={counts['phase1_pass']}  "
                      f"p2_pass={counts['phase2_pass']}")
        print("=" * 72)

    # Write output JSON (same format as input + added keys)
    if output_json:
        with open(output_json, "w", encoding="utf-8") as fw:
            json.dump(reviewed, fw, ensure_ascii=False, indent=4)
        if verbose:
            print(f"\n✅ Wrote review results → {output_json}")

    # Write stats to separate file
    if stats_json:
        with open(stats_json, "w", encoding="utf-8") as fw:
            json.dump(stats, fw, ensure_ascii=False, indent=4)
        if verbose:
            print(f"✅ Wrote statistics → {stats_json}")

    return reviewed


# ---------------------------------------------------------------------------
# Pretty-print helper
# ---------------------------------------------------------------------------


def print_review(review: dict[str, Any]) -> None:
    print("=" * 72)
    print(f"DISTORTION REVIEW – {review.get('xlsx')}")
    print("=" * 72)
    print(f"Q: {review.get('question')}")
    print(f"A: {review.get('answer')}\n")

    for r in review.get("results", []):
        name = r["distortion"]
        verdict = r.get("verdict", "error")
        icon = {
            "validated": "✅",
            "ineffective": "🟡",
            "destructive": "🔴",
        }.get(verdict, "⚠")
        print(f"{icon} {name:30s} → {verdict}")
        direct = r.get("direct", {})
        dd = r.get("dedistort")
        print(f"     direct  : {direct.get('answer')!r:40s} "
              f"correct={direct.get('correct')}")
        if dd is not None:
            print(f"     de-dist : {dd.get('answer')!r:40s} "
                  f"correct={dd.get('correct')}  "
                  f"(detected: {dd.get('detected_distortion')})")
        if "error" in r:
            print(f"     error   : {r['error']}")

    s = review.get("summary", {})
    print("-" * 72)
    print(f"validated={s.get('validated')}/{s.get('total')}  "
          f"ineffective={s.get('ineffective')}  "
          f"destructive={s.get('destructive')}  "
          f"errors={s.get('errors')}")
    print("=" * 72)


# ---------------------------------------------------------------------------
# Script entry-point
# ---------------------------------------------------------------------------


if __name__ == "__main__":
    run_reviewer(
        samples_json="30_Samples.json",
        distorted_dir="Output/04162026",
        output_json="30_Samples_reviewed.json",
        limit=None,
        store_validated=True,
        verbose=True,
        model="gpt-5.2")
