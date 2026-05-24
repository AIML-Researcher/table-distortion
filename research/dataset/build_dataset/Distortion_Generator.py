"""
Distortion Generator
====================
Generates realistic table distortions based on the categories defined in
"Distortions Types.md".  Each distortion is a standalone function that
accepts a pandas DataFrame and returns a distorted DataFrame.

Usage (CLI):
    python Distortion_Generator.py --input data.csv --output distorted.xlsx --distortion all
    python Distortion_Generator.py --input data.xlsx --output out.xlsx --distortion ocr_char_misinterpretation

Usage (Library):
    from Distortion_Generator import ocr_char_misinterpretation
    distorted_df = ocr_char_misinterpretation(df)
"""

import argparse
import json
import os
import random
import re
from pathlib import Path
from typing import Callable
import pandas as pd
from azure.identity import DefaultAzureCredential, get_bearer_token_provider
from openai import AzureOpenAI
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

OCR_CHAR_MAP = {
    "0": "O", "O": "0",
    "1": "l", "l": "1", "I": "1",
    "5": "S", "S": "5",
    "8": "B", "B": "8",
    "2": "Z", "Z": "2",
}

FORMATTING_SYMBOLS = ["%", "$", "€", "£", "¥", "°", "±", "~"]

UNIT_PATTERNS = [
    r"\bkg\b", r"\bkm\b", r"\bcm\b", r"\bmm\b", r"\bmi\b",
    r"\blb\b", r"\boz\b", r"\bft\b", r"\bin\b", r"\bm\b",
    r"\$", r"€", r"£", r"¥", r"%",
]

NOISE_CHARS = list("#@?~^&*!`|")


# ---------------------------------------------------------------------------
# LLM-based target selection
# ---------------------------------------------------------------------------

def _get_llm_client():
    """Create and return an Azure OpenAI client (cached on module level)."""
    endpoint = os.getenv(
        "ENDPOINT_URL",
        "https://oaiscience-oai-swc.openai.azure.com/",
    )
    deployment = os.getenv("DEPLOYMENT_NAME", "gpt-4o-mini")
    token_provider = get_bearer_token_provider(
        DefaultAzureCredential(),
        "https://cognitiveservices.azure.com/.default",
    )
    client = AzureOpenAI(
        azure_endpoint=endpoint,
        azure_ad_token_provider=token_provider,
        api_version="2025-01-01-preview",
    )
    return client, deployment


def _df_to_markdown(df: pd.DataFrame, max_rows: int = 60) -> str:
    """Serialise a DataFrame to a compact Markdown table string."""
    if len(df) > max_rows:
        top = df.head(max_rows // 2)
        bot = df.tail(max_rows // 2)
        ellipsis_row = pd.DataFrame(
            [["…"] * len(df.columns)], columns=df.columns
        )
        df = pd.concat([top, ellipsis_row, bot], ignore_index=True)
    return df.to_markdown(index=False)


def _llm_select_targets(df: pd.DataFrame, question: str, answer: str,
                        system_prompt: str) -> dict:
    """Call the LLM with a tailored prompt and return the parsed JSON response
    containing the selected row/column indices for a specific distortion."""
    client, deployment = _get_llm_client()
    table_md = _df_to_markdown(df)

    user_message = (
        f"### Table\n{table_md}\n\n"
        f"### Question\n{question}\n\n"
        f"### Answer\n{answer}\n\n"
        f"### Table Shape\nRows: {len(df)}, Columns: {len(df.columns)}\n"
        f"### Column Names (with indices)\n"
        + "\n".join(f"  {i}: {col}" for i, col in enumerate(df.columns))
    )

    response = client.chat.completions.create(
        model=deployment,
        temperature=0.2,
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user",   "content": user_message},
        ],
        response_format={"type": "json_object"},
    )

    raw = response.choices[0].message.content.strip()
    if raw.startswith("```"):
        raw = re.sub(r"^```(?:json)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)

    return json.loads(raw)


# --- Tailored system prompts for each distortion ---

_MERGE_CELLS_PROMPT = """\
You are an expert at selecting which table header columns to merge to maximally \
disrupt a QA solver's ability to answer a specific question.

You will receive a table, a question, and the correct answer. Your task is to \
select a contiguous range of header columns to merge such that the merge makes \
it impossible to directly look up the answer column(s) without first \
de-distorting (un-merging) the headers.

**Strategy**: Identify which column headers are critical for answering the \
question (the column the answer comes from, or columns used for filtering/ \
lookup). Merge a range that includes at least one of these critical columns \
with an adjacent column.

Return a JSON object with:
```json
{
  "reasoning": "<brief explanation of why these columns were chosen>",
  "start_col": <0-based start column index>,
  "end_col": <0-based end column index (inclusive)>
}
```
The range must be valid: 0 <= start_col < end_col < num_columns.
Return ONLY the JSON object.
"""

_HORIZONTAL_SHIFT_PROMPT = """\
You are an expert at selecting which table rows to shift horizontally (misalign) to maximally \
disrupt a QA solver's ability to answer a specific question.

You will receive a table, a question, and the correct answer. Your task is to \
select specific data rows where shifting their values one position to the right \
would corrupt the answer or the lookup path to the answer.

**Strategy**: Identify which rows contain the answer or are critical for \
filtering/aggregation to find the answer. Select those rows for misalignment \
so their values end up under wrong column headers.

Return a JSON object with:
```json
{
  "reasoning": "<brief explanation of why these rows were chosen>",
  "row_indices": [<list of 0-based data row indices to shift>]
}
```
Row indices must be valid: 0 <= index < number_of_data_rows.
Select 1-5 rows. Return ONLY the JSON object.
"""

_VERTICAL_SHIFT_PROMPT = """\
You are an expert at selecting which table COLUMNS to shift vertically downward \
to maximally disrupt a QA solver's ability to answer a specific question.

You will receive a table, a question, and the correct answer. Shifting a column \
vertically means its header and every value move down by N rows; the top N \
cells of that column become empty and the table grows by N rows at the bottom. \
This breaks the row-level alignment between the shifted columns and the rest of \
the table.

**Strategy**: Identify the columns that are required to answer the question \
(the column that contains the answer plus any columns used for \
filtering/lookup). Select a SUBSET of those query-relevant columns to shift, \
leaving at least one query-relevant column unshifted so that row alignment \
between query-relevant columns is destroyed.

Return a JSON object with:
```json
{
  "reasoning": "<brief explanation of why these columns were chosen>",
  "col_indices": [<list of 0-based column indices to shift down>]
}
```
Column indices must be valid: 0 <= index < num_columns.
Select 1-4 columns, and ALWAYS leave at least one query-relevant column \
unshifted (unless there is only one query-relevant column in total). \
Return ONLY the JSON object.
"""

_BROKEN_ROWS_SPLIT_PROMPT = """\
You are an expert at selecting which table rows to split in half to maximally \
disrupt a QA solver's ability to answer a specific question.

You will receive a table, a question, and the correct answer. Splitting a row \
means the left half of its columns go on one row and the right half on the next, \
breaking row integrity.

**Strategy**: Identify rows that contain the answer or are essential for \
answering the question (e.g., the row being looked up, or rows needed for \
aggregation). Split those rows so the solver cannot read them as single records.

Return a JSON object with:
```json
{
  "reasoning": "<brief explanation of why these rows were chosen>",
  "row_indices": [<list of 0-based data row indices to split>]
}
```
Row indices must be valid: 0 <= index < number_of_data_rows.
Select 1-3 rows. Return ONLY the JSON object.
"""

_BROKEN_ROWS_MERGE_PROMPT = """\
You are an expert at selecting which pairs of adjacent table rows to merge \
(concatenate) to maximally disrupt a QA solver's ability to answer a question.

You will receive a table, a question, and the correct answer. Merging two \
adjacent rows means their cell values get concatenated, destroying individual \
cell look-ups.

**Strategy**: Identify the row containing the answer or rows critical for \
the lookup. Select a pair where at least one row is answer-relevant, so that \
merging corrupts the answer cell or the filtering condition.

Return a JSON object with:
```json
{
  "reasoning": "<brief explanation of why these row pairs were chosen>",
  "start_indices": [<list of 0-based indices of the FIRST row in each pair>]
}
```
Each start_index pairs with start_index+1. Pairs must not overlap. \
Indices must satisfy: 0 <= start_index < number_of_data_rows - 1.
Select 1-2 pairs. Return ONLY the JSON object.
"""

_MULTI_COLUMN_COLLAPSE_PROMPT = """\
You are an expert at selecting which adjacent columns to collapse (merge into \
one) to maximally disrupt a QA solver's ability to answer a specific question.

You will receive a table, a question, and the correct answer. Collapsing two \
columns means their values are concatenated with a space, and the two original \
columns are replaced by one.

**Strategy**: Identify the column that contains the answer or the column used \
for filtering/lookup. Collapse that column with an adjacent one so the solver \
cannot parse individual values without first splitting the column.

Return a JSON object with:
```json
{
  "reasoning": "<brief explanation of why this column was chosen>",
  "col_index": <0-based index of the LEFT column in the pair to collapse>
}
```
col_index must satisfy: 0 <= col_index < num_columns - 1.
Return ONLY the JSON object.
"""

_FOOTNOTE_INJECTION_PROMPT = """\
You are an expert at selecting where to inject fake footnote rows into a table \
to maximally disrupt a QA solver's ability to answer a specific question.

You will receive a table, a question, and the correct answer. Footnote rows are \
noise rows like "* See appendix for details." inserted into the table body.

**Strategy**: Identify the row positions that are most critical for answering \
the question — near the answer row, or within a range the solver would need to \
scan/aggregate. Insert footnotes at those positions to break row counting, \
interrupt contiguous data ranges, and confuse filtering.

Return a JSON object with:
```json
{
  "reasoning": "<brief explanation of why these positions were chosen>",
  "positions": [<list of 1-based row positions where footnotes should be inserted>]
}
```
Positions must be valid: 1 <= position <= number_of_data_rows.
Select 1-3 positions. Return ONLY the JSON object.
"""


# ---------------------------------------------------------------------------
# Helper utilities
# ---------------------------------------------------------------------------


def _load_input(filepath: str) -> pd.DataFrame:
    """Load a CSV or Excel file into a DataFrame."""
    p = Path(filepath)
    if p.suffix.lower() == ".csv":
        return pd.read_csv(filepath, dtype=str, keep_default_na=False)
    elif p.suffix.lower() in (".xls", ".xlsx"):
        return pd.read_excel(filepath, dtype=str, keep_default_na=False)
    else:
        raise ValueError(f"Unsupported file format: {p.suffix}")


def _save_single_sheet(df: pd.DataFrame, filepath: str, sheet_name: str = "Distorted"):
    """Save a single DataFrame to an Excel file."""
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)


def _df_to_workbook(df: pd.DataFrame, sheet_name: str = "Sheet1") -> Workbook:
    """Convert a DataFrame to an openpyxl Workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append([str(c) if c is not None else "" for c in r])
    return wb


# =========================================================================
# 1. STRUCTURAL DISTORTIONS (Layout Breaks)
# =========================================================================


def merge_cells(df: pd.DataFrame, seed: int | None = None,
                question: str | None = None, answer: str | None = None) -> Workbook:
    """Merge adjacent header cells in the output Excel workbook.

    When question/answer are provided, uses LLM to select which columns
    to merge for maximum impact on answerability. Falls back to random
    selection otherwise.

    Returns an openpyxl Workbook (not a DataFrame) because merged cells
    are a formatting concept that only exists in Excel.
    """
    wb = _df_to_workbook(df)
    ws = wb.active
    num_cols = ws.max_column
    if num_cols < 2:
        return wb

    if question and answer:
        try:
            result = _llm_select_targets(df, question, answer, _MERGE_CELLS_PROMPT)
            start_col = max(1, result["start_col"] + 1)  # convert 0-based to 1-based
            end_col = min(num_cols, result["end_col"] + 1)
            if start_col >= end_col:
                end_col = min(start_col + 1, num_cols)
            # print(f"    LLM selected merge range: cols {start_col}-{end_col} | {result.get('reasoning', '')}")
        except Exception as e:
            print(f"    ⚠ LLM selection failed ({e}), falling back to random")
            rng = random.Random(seed)
            start_col = rng.randint(1, max(1, num_cols - 1))
            end_col = rng.randint(start_col + 1, num_cols)
    else:
        rng = random.Random(seed)
        start_col = rng.randint(1, max(1, num_cols - 1))
        end_col = rng.randint(start_col + 1, num_cols)

    ws.merge_cells(start_row=1, start_column=start_col,
                   end_row=1, end_column=end_col)
    return wb


def horizontal_shift(df: pd.DataFrame, n_rows: int = 5,
                     shift: int = 1, seed: int | None = None,
                     question: str | None = None, answer: str | None = None) -> pd.DataFrame:
    """Shift data in selected rows to the right by *shift* positions,
    simulating column misalignment from faulty extraction. Previously named
    ``column_misalignment``.

    When question/answer are provided, uses LLM to select which rows
    to shift for maximum impact on answerability.
    """
    df = df.copy()
    n_rows = min(n_rows, len(df))

    if question and answer:
        try:
            result = _llm_select_targets(df, question, answer, _HORIZONTAL_SHIFT_PROMPT)
            rows_to_shift = [i for i in result["row_indices"] if 0 <= i < len(df)]
            if not rows_to_shift:
                raise ValueError("LLM returned no valid row indices")
            # print(f"    LLM selected rows to shift: {rows_to_shift} | {result.get('reasoning', '')}")
        except Exception as e:
            print(f"    ⚠ LLM selection failed ({e}), falling back to random")
            rng = random.Random(seed)
            rows_to_shift = rng.sample(range(len(df)), k=n_rows)
    else:
        rng = random.Random(seed)
        rows_to_shift = rng.sample(range(len(df)), k=n_rows)

    # Add extra columns to accommodate the overflow from shifting
    original_col_count = len(df.columns)
    for i in range(shift):
        new_col_name = f" "
        df[new_col_name] = ""

    for idx in rows_to_shift:
        values = df.iloc[idx].tolist()[:original_col_count]
        shifted = [""] * shift + values
        df.iloc[idx] = shifted

    return df


def broken_rows_split(df: pd.DataFrame, n_rows: int = 4,
                      seed: int | None = None,
                      question: str | None = None, answer: str | None = None) -> pd.DataFrame:
    """Split selected rows into two rows each – first row gets the left
    half of columns, second row gets the right half.

    When question/answer are provided, uses LLM to select which rows
    to split for maximum impact on answerability.
    """
    df = df.copy().reset_index(drop=True)
    n_rows = min(n_rows, len(df))
    mid = len(df.columns) // 2

    if question and answer:
        try:
            result = _llm_select_targets(df, question, answer, _BROKEN_ROWS_SPLIT_PROMPT)
            rows_to_split = sorted(
                [i for i in result["row_indices"] if 0 <= i < len(df)],
                reverse=True
            )
            if not rows_to_split:
                raise ValueError("LLM returned no valid row indices")
            # print(f"    LLM selected rows to split: {rows_to_split} | {result.get('reasoning', '')}")
        except Exception as e:
            print(f"    ⚠ LLM selection failed ({e}), falling back to random")
            rng = random.Random(seed)
            rows_to_split = sorted(rng.sample(range(len(df)), k=n_rows), reverse=True)
    else:
        rng = random.Random(seed)
        rows_to_split = sorted(rng.sample(range(len(df)), k=n_rows), reverse=True)

    for idx in rows_to_split:
        original = df.iloc[idx].tolist()
        row_top = original[:mid] + [""] * (len(df.columns) - mid)
        row_bot = [""] * mid + original[mid:]

        top_series = pd.DataFrame([row_top], columns=df.columns)
        bot_series = pd.DataFrame([row_bot], columns=df.columns)

        df = pd.concat([
            df.iloc[:idx],
            top_series,
            bot_series,
            df.iloc[idx + 1:]
        ], ignore_index=True)

    return df


def broken_rows_merge(df: pd.DataFrame, n_pairs: int = 1,
                      seed: int | None = None,
                      question: str | None = None, answer: str | None = None) -> pd.DataFrame:
    """Merge pairs of adjacent rows by concatenating their cell values.

    When question/answer are provided, uses LLM to select which row
    pairs to merge for maximum impact on answerability.
    """
    df = df.copy().reset_index(drop=True)
    if len(df) < 2:
        return df

    n_pairs = min(n_pairs, len(df) // 2)

    if question and answer:
        try:
            result = _llm_select_targets(df, question, answer, _BROKEN_ROWS_MERGE_PROMPT)
            start_indices = sorted(
                [i for i in result["start_indices"] if 0 <= i < len(df) - 1],
                reverse=True
            )
            if not start_indices:
                raise ValueError("LLM returned no valid start indices")
            # print(f"    LLM selected row pairs starting at: {start_indices} | {result.get('reasoning', '')}")
        except Exception as e:
            print(f"    ⚠ LLM selection failed ({e}), falling back to random")
            rng = random.Random(seed)
            start_indices = sorted(
                rng.sample(range(0, len(df) - 1, 2), k=min(n_pairs, len(df) // 2)),
                reverse=True,
            )
    else:
        rng = random.Random(seed)
        start_indices = sorted(
            rng.sample(range(0, len(df) - 1, 2), k=min(n_pairs, len(df) // 2)),
            reverse=True,
        )

    for idx in start_indices:
        merged = []
        for col in df.columns:
            v1 = str(df.at[idx, col]).strip()
            v2 = str(df.at[idx + 1, col]).strip()
            merged.append(f"{v1} {v2}".strip())
        merged_row = pd.DataFrame([merged], columns=df.columns)
        df = pd.concat([
            df.iloc[:idx],
            merged_row,
            df.iloc[idx + 2:]
        ], ignore_index=True)

    return df


# =========================================================================
# 2. TEXT RECOGNITION ERRORS (OCR Issues)
# =========================================================================


def ocr_char_misinterpretation(df: pd.DataFrame, probability: float = 0.3,
                               seed: int | None = None) -> pd.DataFrame:
    """Replace characters with common OCR mis-reads (0↔O, 1↔l, 5↔S, …)."""
    rng = random.Random(seed)
    df = df.copy()

    def _corrupt(val: str) -> str:
        chars = list(val)
        for i, ch in enumerate(chars):
            if ch in OCR_CHAR_MAP and rng.random() < probability:
                chars[i] = OCR_CHAR_MAP[ch]
        return "".join(chars)

    for col in df.columns:
        df[col] = df[col].astype(str).apply(_corrupt)

    return df


def ocr_lost_formatting(df: pd.DataFrame) -> pd.DataFrame:
    """Strip formatting symbols (%, $, €, …) from all cells,
    simulating OCR tools that fail to recognise special characters.
    """
    df = df.copy()
    pattern = "[" + re.escape("".join(FORMATTING_SYMBOLS)) + "]"

    for col in df.columns:
        df[col] = df[col].astype(str).apply(lambda v: re.sub(pattern, "", v))

    return df


# =========================================================================
# 3. DATA TYPE DISTORTIONS
# =========================================================================


def numbers_as_text(df: pd.DataFrame) -> pd.DataFrame:
    """Prefix numeric cells with an apostrophe so Excel treats them as text."""
    df = df.copy()

    def _prefix(val: str) -> str:
        stripped = val.strip().replace(",", "")
        try:
            float(stripped)
            return f"'{val}'"          # apostrophe prefix → Excel text
        except ValueError:
            return val

    for col in df.columns:
        df[col] = df[col].astype(str).apply(_prefix)

    return df


def date_format_corruption(df: pd.DataFrame,
                           seed: int | None = None) -> pd.DataFrame:
    """Detect date-like strings and rewrite them in a randomly different
    format, introducing ambiguity (e.g. MM/DD ↔ DD/MM, or to serial
    number).
    """
    rng = random.Random(seed)
    df = df.copy()
    date_re = re.compile(
        r"^(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2,4})$"
    )

    def _corrupt_date(val: str) -> str:
        m = date_re.match(val.strip())
        if not m:
            return val
        a, b, c = m.group(1), m.group(2), m.group(3)
        fmt = rng.choice(["swap", "dash", "serial", "dot"])
        if fmt == "swap":
            return f"{b}/{a}/{c}"          # DD/MM ↔ MM/DD ambiguity
        elif fmt == "dash":
            return f"{c}-{a}-{b}"          # YYYY-MM-DD style
        elif fmt == "dot":
            return f"{a}.{b}.{c}"
        else:  # serial
            try:
                from datetime import datetime
                dt = datetime(int(c), int(a), int(b))
                delta = dt - datetime(1900, 1, 1)
                return str(delta.days + 2)  # Excel serial offset
            except Exception:
                return val

    for col in df.columns:
        df[col] = df[col].astype(str).apply(_corrupt_date)

    return df


def decimal_separator_swap(df: pd.DataFrame) -> pd.DataFrame:
    """Swap '.' and ',' in numeric-looking cells, simulating a locale
    mismatch (e.g. 1,200.50 → 1.200,50).
    """
    df = df.copy()
    num_re = re.compile(r"^[\d.,]+$")

    def _swap(val: str) -> str:
        if num_re.match(val.strip()) and ("." in val or "," in val):
            return val.translate(str.maketrans(".,", ",."))
        return val

    for col in df.columns:
        df[col] = df[col].astype(str).apply(_swap)

    return df


# =========================================================================
# 4. HEADER & SCHEMA PROBLEMS
# =========================================================================


def header_as_data_row(df: pd.DataFrame) -> pd.DataFrame:
    """Push the real header into the first data row and replace the
    header with generic column names (Col_0, Col_1, …).
    """
    header_row = pd.DataFrame([df.columns.tolist()], columns=df.columns)
    df = pd.concat([header_row, df], ignore_index=True)
    df.columns = [f"Col_{i}" for i in range(len(df.columns))]
    return df


# =========================================================================
# 5. COMPLEX TABLE LAYOUT FAILURES
# =========================================================================


def multi_column_collapse(df: pd.DataFrame, col_index: int | None = None,
                          seed: int | None = None,
                          question: str | None = None, answer: str | None = None) -> pd.DataFrame:
    """Merge two adjacent columns into one (space-separated),
    simulating a column-boundary detection failure.

    When question/answer are provided, uses LLM to select which column
    pair to collapse for maximum impact on answerability.
    """
    df = df.copy()
    if len(df.columns) < 2:
        return df

    if col_index is None:
        if question and answer:
            try:
                result = _llm_select_targets(df, question, answer, _MULTI_COLUMN_COLLAPSE_PROMPT)
                col_index = result["col_index"]
                if not (0 <= col_index < len(df.columns) - 1):
                    raise ValueError(f"col_index {col_index} out of range")
                # print(f"    LLM selected column to collapse: {col_index} ({df.columns[col_index]} + {df.columns[col_index+1]}) | {result.get('reasoning', '')}")
            except Exception as e:
                print(f"    ⚠ LLM selection failed ({e}), falling back to random")
                rng = random.Random(seed)
                col_index = rng.randint(0, len(df.columns) - 2)
        else:
            rng = random.Random(seed)
            col_index = rng.randint(0, len(df.columns) - 2)

    col_a = df.columns[col_index]
    col_b = df.columns[col_index + 1]

    merged_name = f"{col_a} {col_b}"
    merged_values = (df[col_a].astype(str) + " " + df[col_b].astype(str)).str.strip()

    df.insert(col_index, merged_name, merged_values)
    df = df.drop(columns=[col_a, col_b])

    return df


def footnote_injection(df: pd.DataFrame, n_footnotes: int = 2,
                       seed: int | None = None,
                       question: str | None = None, answer: str | None = None) -> pd.DataFrame:
    """Inject fake footnote / side-note rows into the table body.

    When question/answer are provided, uses LLM to select positions
    for maximum impact on answerability. Footnote text is still chosen
    randomly.
    """
    rng = random.Random(seed)
    footnotes = [
        "* See appendix for details.",
        "Source: internal report 2024.",
        "Note: values are approximate.",
        "** Revised figures.",
        "† Data unavailable for this period.",
        "1) Figures rounded to nearest integer.",
    ]

    if question and answer:
        try:
            result = _llm_select_targets(df, question, answer, _FOOTNOTE_INJECTION_PROMPT)
            positions = [p for p in result["positions"] if 1 <= p <= len(df)]
            if not positions:
                raise ValueError("LLM returned no valid positions")
            n_footnotes = len(positions)
            # print(f"    LLM selected footnote positions: {positions} | {result.get('reasoning', '')}")
        except Exception as e:
            print(f"    ⚠ LLM selection failed ({e}), falling back to random")
            positions = [rng.randint(1, len(df)) for _ in range(n_footnotes)]
    else:
        positions = [rng.randint(1, len(df)) for _ in range(n_footnotes)]

    # Sort positions in reverse so earlier insertions don't shift later ones
    for pos in sorted(positions, reverse=True):
        note = rng.choice(footnotes)
        row = [note] + [""] * (len(df.columns) - 1)
        top = df.iloc[:pos]
        bottom = df.iloc[pos:]
        note_df = pd.DataFrame([row], columns=df.columns)
        df = pd.concat([top, note_df, bottom], ignore_index=True)

    return df


# =========================================================================
# 6. IMAGE QUALITY–DRIVEN ISSUES
# =========================================================================


def random_noise_characters(df: pd.DataFrame, probability: float = 0.1,
                            seed: int | None = None) -> pd.DataFrame:
    """Insert random junk characters into numerical cells, simulating artefacts
    from low-resolution or noisy scans.
    """
    rng = random.Random(seed)
    df = df.copy()

    def _is_numeric(val: str) -> bool:
        try:
            float(val.strip().replace(",", ""))
            return True
        except ValueError:
            return False

    def _add_noise(val: str) -> str:
        if not val or not _is_numeric(val) or rng.random() > probability:
            return val
        pos = rng.randint(0, max(0, len(val) - 1))
        noise = rng.choice(NOISE_CHARS)
        return val[:pos] + noise + val[pos:]

    for col in df.columns:
        df[col] = df[col].astype(str).apply(_add_noise)

    return df


# =========================================================================
# 7. SEMANTIC DISTORTIONS (Harder to Detect)
# =========================================================================


def context_loss(df: pd.DataFrame) -> pd.DataFrame:
    """Strip unit strings ($, kg, %, …) from values so they lose
    their meaning.  Only targets values that *contain* a number,
    unlike ocr_lost_formatting which strips symbols everywhere.
    """
    df = df.copy()
    combined_pattern = "|".join(UNIT_PATTERNS)

    def _strip_units(val: str) -> str:
        if re.search(r"\d", val):           # only touch numeric-ish cells
            return re.sub(combined_pattern, "", val).strip()
        return val

    for col in df.columns:
        df[col] = df[col].astype(str).apply(_strip_units)

    return df


# =========================================================================
# 8. VERTICAL SHIFT (Column-wise misalignment) – ported from
#    create_vertical_shift_distortions.py
# =========================================================================

_VERTICAL_SHIFT_AMOUNT_MIN = 1
_VERTICAL_SHIFT_AMOUNT_MAX = 4
_VERTICAL_SHIFT_MAX_COLS = 4


def _shift_columns_down(ws, columns_1based: list[int], shift: int) -> None:
    """Shift each given column (entire column including header) down by
    ``shift`` rows in an openpyxl worksheet. The top ``shift`` cells of those
    columns become empty and the worksheet is extended at the bottom so no
    data is lost.
    """
    if shift <= 0 or not columns_1based:
        return
    original_max_row = ws.max_row
    new_max_row = original_max_row + shift
    for col in columns_1based:
        values = [ws.cell(row=r, column=col).value for r in range(1, original_max_row + 1)]
        for r in range(1, new_max_row + 1):
            ws.cell(row=r, column=col).value = None
        for i, v in enumerate(values):
            ws.cell(row=i + 1 + shift, column=col).value = v


def vertical_shift(df: pd.DataFrame, n_cols: int | None = None,
                   shift: int | None = None, seed: int | None = None,
                   question: str | None = None, answer: str | None = None) -> Workbook:
    """Shift one or more entire columns (header + values) downward by N rows.

    The top ``N`` cells of each shifted column become empty and the table
    grows by ``N`` rows at the bottom – destroying the row-level alignment
    between the shifted columns and the rest of the table.

    When ``question``/``answer`` are provided, the LLM is asked to choose
    query-relevant columns to shift (always leaving at least one
    query-relevant column unshifted when possible).

    Returns an openpyxl Workbook because the operation produces empty header
    cells, which cannot be cleanly expressed in a DataFrame.
    """
    wb = _df_to_workbook(df)
    ws = wb.active
    num_cols = ws.max_column
    if num_cols < 1:
        return wb

    rng = random.Random(seed)

    cols_0based: list[int] = []
    if question and answer:
        try:
            result = _llm_select_targets(df, question, answer, _VERTICAL_SHIFT_PROMPT)
            raw_cols = result.get("col_indices", [])
            cols_0based = [int(c) for c in raw_cols if 0 <= int(c) < num_cols]
            if not cols_0based:
                raise ValueError("LLM returned no valid col indices")
            # Cap to MAX_COLS
            cols_0based = cols_0based[:_VERTICAL_SHIFT_MAX_COLS]
            # print(f"    LLM selected vertical_shift cols: {cols_0based} | {result.get('reasoning', '')}")
        except Exception as e:
            print(f"    ⚠ LLM selection failed ({e}), falling back to random")
            cols_0based = []

    if not cols_0based:
        max_pick = n_cols if n_cols else rng.randint(1, min(num_cols, _VERTICAL_SHIFT_MAX_COLS))
        max_pick = max(1, min(max_pick, num_cols))
        cols_0based = rng.sample(range(num_cols), max_pick)

    if shift is None or shift <= 0:
        shift = rng.randint(_VERTICAL_SHIFT_AMOUNT_MIN, _VERTICAL_SHIFT_AMOUNT_MAX)

    cols_1based = [c + 1 for c in cols_0based]
    _shift_columns_down(ws, cols_1based, shift)
    return wb


# Backward-compatibility alias – the function was renamed from
# ``column_misalignment`` to ``horizontal_shift``.
column_misalignment = horizontal_shift


# =========================================================================
# Registry of all distortion functions
# =========================================================================

# Functions that accept question/answer for LLM-guided selection
LLM_GUIDED_DISTORTIONS = {
    "merge_cells", "horizontal_shift", "broken_rows_split",
    "broken_rows_merge", "multi_column_collapse", "footnote_injection",
    "vertical_shift",
}

DISTORTIONS: dict[str, Callable] = {
    # 1. Structural
    "merge_cells":               merge_cells,
    "horizontal_shift":          horizontal_shift,
    "vertical_shift":            vertical_shift,
    "broken_rows_split":         broken_rows_split,
    "broken_rows_merge":         broken_rows_merge,
    # 2. OCR
    "ocr_char_misinterpret":     ocr_char_misinterpretation,
    "ocr_lost_formatting":       ocr_lost_formatting,
    # 3. Data Type
    "numbers_as_text":           numbers_as_text,
    "date_format_corruption":    date_format_corruption,
    "decimal_separator_swap":    decimal_separator_swap,
    # 4. Header & Schema
    "header_as_data_row":        header_as_data_row,
    # 5. Complex Layout
    "multi_column_collapse":     multi_column_collapse,
    "footnote_injection":        footnote_injection,
    # 6. Image Quality
    "random_noise_chars":        random_noise_characters,
    # 7. Semantic
    "context_loss":              context_loss
}


# =========================================================================
# apply_selected_distortions  –  one sheet per distortion in a single .xlsx
# =========================================================================


def apply_selected_distortions(df: pd.DataFrame, output_path: str,
                               distortion_names: list[str],
                               seed: int | None = None,
                               question: str | None = None,
                               answer: str | None = None) -> None:
    """Apply a list of distortions and write each result as a
    separate sheet in a single .xlsx workbook.
    """
    wb = Workbook()
    wb.remove(wb.active)

    ws_orig = wb.create_sheet(title="original")
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_orig.append(list(r))

    for name in distortion_names:
        func = DISTORTIONS[name]
        print(f"  Applying: {name}")
        try:
            kwargs = {}
            if "seed" in func.__code__.co_varnames:
                kwargs["seed"] = seed
            if name in LLM_GUIDED_DISTORTIONS and question and answer:
                kwargs["question"] = question
                kwargs["answer"] = answer
            result = func(df, **kwargs)
        except Exception as e:
            print(f"    ⚠ Skipped {name}: {e}")
            continue

        if isinstance(result, Workbook):
            src_ws = result.active
            ws = wb.create_sheet(title=name[:31])
            for row in src_ws.iter_rows(values_only=False):
                ws.append([cell.value for cell in row])
            for merge_range in src_ws.merged_cells.ranges:
                ws.merge_cells(str(merge_range))
        else:
            ws = wb.create_sheet(title=name[:31])
            for r in dataframe_to_rows(result, index=False, header=True):
                ws.append(list(r))

    wb.save(output_path)
    print(f"\n✅ Saved distortions → {output_path}")


# =========================================================================
# apply_all_distortions  –  one sheet per distortion in a single .xlsx
# =========================================================================


def apply_all_distortions(df: pd.DataFrame, output_path: str,
                          seed: int | None = None,
                          question: str | None = None,
                          answer: str | None = None) -> None:
    """Apply every registered distortion and write each result as a
    separate sheet in a single .xlsx workbook.
    """
    wb = Workbook()
    wb.remove(wb.active)

    ws_orig = wb.create_sheet(title="original")
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_orig.append(list(r))

    for name, func in DISTORTIONS.items():
        print(f"  Applying: {name}")
        try:
            kwargs = {}
            if "seed" in func.__code__.co_varnames:
                kwargs["seed"] = seed
            if name in LLM_GUIDED_DISTORTIONS and question and answer:
                kwargs["question"] = question
                kwargs["answer"] = answer
            result = func(df, **kwargs)
        except Exception as e:
            print(f"    ⚠ Skipped {name}: {e}")
            continue

        if isinstance(result, Workbook):
            src_ws = result.active
            ws = wb.create_sheet(title=name[:31])
            for row in src_ws.iter_rows(values_only=False):
                ws.append([cell.value for cell in row])
            for merge_range in src_ws.merged_cells.ranges:
                ws.merge_cells(str(merge_range))
        else:
            ws = wb.create_sheet(title=name[:31])
            for r in dataframe_to_rows(result, index=False, header=True):
                ws.append(list(r))

    wb.save(output_path)
    print(f"\n✅ Saved all distortions → {output_path}")


# =========================================================================
# CLI entry point
# =========================================================================


def main():

    parser = argparse.ArgumentParser(
        description="Generate distorted tables from a CSV / Excel file."
    )
    parser.add_argument(
        "--input", "-i", required=True,
        help="Path to the input CSV or Excel file.",
    )
    parser.add_argument(
        "--output", "-o", default="distorted_output.xlsx",
        help="Path for the output .xlsx file (default: distorted_output.xlsx).",
    )
    parser.add_argument(
        "--distortion", "-d", nargs="+", default=["all"],
        help="One or more distortion names to apply, or 'all' (default).",
    )
    parser.add_argument(
        "--seed", "-s", type=int, default=42,
        help="Random seed for reproducibility.",
    )
    parser.add_argument(
        "--question", "-q", default=None,
        help="The question for LLM-guided distortion targeting.",
    )
    parser.add_argument(
        "--answer", "-a", default=None,
        help="The answer for LLM-guided distortion targeting.",
    )
    args = parser.parse_args()

    # Validate distortion names
    for d in args.distortion:
        if d != "all" and d not in DISTORTIONS:
            parser.error(f"Unknown distortion: '{d}'. Choose from: {', '.join(DISTORTIONS.keys())}, all")

    # Load input
    df = _load_input(args.input)
    print(f"Loaded {len(df)} rows × {len(df.columns)} cols from '{args.input}'\n")

    if "all" in args.distortion:
        apply_all_distortions(df, args.output, seed=args.seed,
                              question=args.question, answer=args.answer)
    elif len(args.distortion) == 1:
        name = args.distortion[0]
        func = DISTORTIONS[name]
        kwargs = {}
        if "seed" in func.__code__.co_varnames:
            kwargs["seed"] = args.seed
        if name in LLM_GUIDED_DISTORTIONS and args.question and args.answer:
            kwargs["question"] = args.question
            kwargs["answer"] = args.answer
        result = func(df, **kwargs)

        if isinstance(result, Workbook):
            result.save(args.output)
        else:
            _save_single_sheet(result, args.output, sheet_name=name[:31])

        print(f"✅ Applied '{name}' → {args.output}")
    else:
        apply_selected_distortions(df, args.output, args.distortion, seed=args.seed,
                                   question=args.question, answer=args.answer)


def gen_distortions(input_path: str, output_path: str, distortion="all", seed=42,
                    question: str | None = None, answer: str | None = None,
                    individual_workbook: bool = False) -> None:

    df = _load_input(input_path)
    print(f"Loaded {len(df)} rows × {len(df.columns)} cols from '{input_path}'\n")

    if individual_workbook:
        base, ext = os.path.splitext(output_path)
        if not ext:
            ext = ".xlsx"

        # Save original
        orig_path = f"{base}_original{ext}"
        _save_single_sheet(df, orig_path, sheet_name="original")
        print(f"✅ Saved original → {orig_path}")

        # Determine distortion list
        if distortion == "all":
            distortion_list = list(DISTORTIONS.keys())
        elif isinstance(distortion, list):
            distortion_list = distortion
        else:
            distortion_list = [distortion]

        for name in distortion_list:
            if name not in DISTORTIONS:
                raise ValueError(f"Unknown distortion: '{name}'. Choose from: {', '.join(DISTORTIONS.keys())}")
            func = DISTORTIONS[name]
            print(f"  Applying: {name}")
            try:
                kwargs = {}
                if "seed" in func.__code__.co_varnames:
                    kwargs["seed"] = seed
                if name in LLM_GUIDED_DISTORTIONS and question and answer:
                    kwargs["question"] = question
                    kwargs["answer"] = answer
                result = func(df, **kwargs)
            except Exception as e:
                print(f"    ⚠ Skipped {name}: {e}")
                continue

            dist_path = f"{base}_{name}{ext}"
            if isinstance(result, Workbook):
                result.save(dist_path)
            else:
                _save_single_sheet(result, dist_path, sheet_name=name[:31])
            print(f"✅ Saved '{name}' → {dist_path}")

        return

    if isinstance(distortion, list):
        for d in distortion:
            if d not in DISTORTIONS:
                raise ValueError(f"Unknown distortion: '{d}'. Choose from: {', '.join(DISTORTIONS.keys())}")
        if len(distortion) == 1:
            name = distortion[0]
            func = DISTORTIONS[name]
            kwargs = {}
            if "seed" in func.__code__.co_varnames:
                kwargs["seed"] = seed
            if name in LLM_GUIDED_DISTORTIONS and question and answer:
                kwargs["question"] = question
                kwargs["answer"] = answer
            result = func(df, **kwargs)

            if isinstance(result, Workbook):
                result.save(output_path)
            else:
                _save_single_sheet(result, output_path, sheet_name=name[:31])

            print(f"✅ Applied '{name}' → {output_path}")
        else:
            apply_selected_distortions(df, output_path, distortion, seed=seed,
                                       question=question, answer=answer)
    elif distortion == "all":
        apply_all_distortions(df, output_path, seed=seed,
                              question=question, answer=answer)
    else:
        func = DISTORTIONS[distortion]
        kwargs = {}
        if "seed" in func.__code__.co_varnames:
            kwargs["seed"] = seed
        if distortion in LLM_GUIDED_DISTORTIONS and question and answer:
            kwargs["question"] = question
            kwargs["answer"] = answer
        result = func(df, **kwargs)

        if isinstance(result, Workbook):
            result.save(output_path)
        else:
            _save_single_sheet(result, output_path, sheet_name=distortion[:31])

        print(f"✅ Applied '{distortion}' → {output_path}")


if __name__ == "__main__":
    # main()
    gen_distortions(input_path="csv/200-csv/3.csv", output_path="test_distorted_output.xlsx", distortion="all", seed=42)
